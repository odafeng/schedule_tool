"""
Excel 日曆形式匯出器
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar
from datetime import datetime
from typing import Dict, List
from backend.models import Doctor, ScheduleSlot


class ExcelCalendarExporter:
    """Excel 日曆形式匯出器"""
    
    def __init__(self, schedule: Dict[str, ScheduleSlot], 
                 doctors: List[Doctor],
                 weekdays: List[str], 
                 holidays: List[str],
                 year: int, 
                 month: int):
        self.schedule = schedule
        self.doctors = doctors
        self.weekdays = weekdays
        self.holidays = holidays
        self.year = year
        self.month = month
        
        # 樣式定義
        self.styles = self._define_styles()
    
    def _define_styles(self):
        """定義 Excel 樣式"""
        return {
            'title': {
                'font': Font(size=16, bold=True),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'font_color': Font(color='FFFFFF', size=16, bold=True)
            },
            'weekday_header': {
                'font': Font(size=12, bold=True, color='FFFFFF'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            },
            'date_cell': {
                'font': Font(size=11, bold=True),
                'alignment': Alignment(horizontal='left', vertical='top')
            },
            'holiday_cell': {
                'fill': PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            },
            'weekend_cell': {
                'fill': PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')
            },
            'weekday_cell': {
                'fill': PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
            },
            'attending_text': {
                'font': Font(size=10, color='0F5132'),
                'alignment': Alignment(horizontal='left', wrap_text=True)
            },
            'resident_text': {
                'font': Font(size=10, color='084C61'),
                'alignment': Alignment(horizontal='left', wrap_text=True)
            },
            'empty_slot': {
                'font': Font(size=10, color='CC0000', bold=True),
                'alignment': Alignment(horizontal='center')
            },
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def export(self, filename: str):
        """匯出為 Excel 檔案"""
        wb = Workbook()
        
        # 移除預設的工作表
        wb.remove(wb.active)
        
        # 創建月曆工作表
        self._create_calendar_sheet(wb)
        
        # 創建統計工作表
        self._create_statistics_sheet(wb)
        
        # 創建醫師清單工作表
        self._create_doctors_sheet(wb)
        
        # 儲存檔案
        wb.save(filename)
        return filename
    
    def _create_calendar_sheet(self, wb):
        """創建月曆形式的工作表"""
        ws = wb.create_sheet(title=f"{self.year}年{self.month}月排班表")
        
        # 設定欄寬
        for i in range(1, 8):
            ws.column_dimensions[get_column_letter(i)].width = 20
        
        # 標題
        ws.merge_cells('A1:G1')
        ws['A1'] = f"{self.year}年{self.month}月 醫師排班表"
        ws['A1'].font = self.styles['title']['font_color']
        ws['A1'].alignment = self.styles['title']['alignment']
        ws['A1'].fill = self.styles['title']['fill']
        
        # 星期標題
        weekday_names = ['週一', '週二', '週三', '週四', '週五', '週六', '週日']
        for i, day_name in enumerate(weekday_names, 1):
            cell = ws.cell(row=2, column=i, value=day_name)
            cell.font = self.styles['weekday_header']['font']
            cell.alignment = self.styles['weekday_header']['alignment']
            cell.fill = self.styles['weekday_header']['fill']
            cell.border = self.styles['border']
        
        # 生成月曆
        cal = calendar.monthcalendar(self.year, self.month)
        current_row = 3
        
        for week in cal:
            # 設定行高
            ws.row_dimensions[current_row].height = 60
            
            for day_of_week, day in enumerate(week, 1):
                if day == 0:
                    # 空白格
                    cell = ws.cell(row=current_row, column=day_of_week)
                    cell.border = self.styles['border']
                else:
                    date_str = f"{self.year:04d}-{self.month:02d}-{day:02d}"
                    cell = ws.cell(row=current_row, column=day_of_week)
                    
                    # 判斷日期類型並設定背景色
                    if date_str in self.holidays:
                        cell.fill = self.styles['holiday_cell']['fill']
                    elif day_of_week in [6, 7]:  # 週末
                        cell.fill = self.styles['weekend_cell']['fill']
                    else:
                        cell.fill = self.styles['weekday_cell']['fill']
                    
                    # 填入日期和排班資訊
                    cell_content = f"{day}日\n"
                    
                    if date_str in self.schedule:
                        slot = self.schedule[date_str]
                        
                        if slot.attending:
                            cell_content += f"主治: {slot.attending}\n"
                        else:
                            cell_content += "主治: ❌未排\n"
                        
                        if slot.resident:
                            cell_content += f"總醫: {slot.resident}"
                        else:
                            cell_content += "總醫: ❌未排"
                    else:
                        cell_content += "未排班"
                    
                    cell.value = cell_content
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    cell.border = self.styles['border']
                    
                    # 設定字體顏色
                    if "❌" in cell_content:
                        cell.font = Font(size=10, color='CC0000')
                    else:
                        cell.font = Font(size=10)
            
            current_row += 1
        
        # 添加圖例
        legend_row = current_row + 2
        ws[f'A{legend_row}'] = "圖例說明："
        ws[f'A{legend_row}'].font = Font(bold=True)
        
        legend_items = [
            ('B', '假日', 'FFE6E6'),
            ('C', '週末', 'FFF4E6'),
            ('D', '平日', 'E6F3FF'),
            ('E', '❌未排', None)
        ]
        
        for col, label, color in legend_items:
            cell = ws[f'{col}{legend_row}']
            cell.value = label
            if color:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.border = self.styles['border']
            cell.alignment = Alignment(horizontal='center')
    
    def _create_statistics_sheet(self, wb):
        """創建統計工作表"""
        ws = wb.create_sheet(title="統計分析")
        
        # 統計每個醫師的值班數
        doctor_stats = {}
        for doctor in self.doctors:
            doctor_stats[doctor.name] = {
                '角色': doctor.role,
                '平日值班': 0,
                '假日值班': 0,
                '總值班數': 0,
                '平日配額': doctor.weekday_quota,
                '假日配額': doctor.holiday_quota
            }
        
        # 計算統計
        for date_str, slot in self.schedule.items():
            is_holiday = date_str in self.holidays
            
            if slot.attending and slot.attending in doctor_stats:
                if is_holiday:
                    doctor_stats[slot.attending]['假日值班'] += 1
                else:
                    doctor_stats[slot.attending]['平日值班'] += 1
                doctor_stats[slot.attending]['總值班數'] += 1
            
            if slot.resident and slot.resident in doctor_stats:
                if is_holiday:
                    doctor_stats[slot.resident]['假日值班'] += 1
                else:
                    doctor_stats[slot.resident]['平日值班'] += 1
                doctor_stats[slot.resident]['總值班數'] += 1
        
        # 寫入表頭
        headers = ['醫師姓名', '角色', '平日值班', '平日配額', '平日使用率', 
                   '假日值班', '假日配額', '假日使用率', '總值班數']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.styles['border']
        
        # 寫入資料
        row = 2
        for doctor_name, stats in doctor_stats.items():
            ws.cell(row=row, column=1, value=doctor_name).border = self.styles['border']
            ws.cell(row=row, column=2, value=stats['角色']).border = self.styles['border']
            ws.cell(row=row, column=3, value=stats['平日值班']).border = self.styles['border']
            ws.cell(row=row, column=4, value=stats['平日配額']).border = self.styles['border']
            
            # 平日使用率
            weekday_rate = stats['平日值班'] / max(stats['平日配額'], 1)
            cell = ws.cell(row=row, column=5, value=f"{weekday_rate:.1%}")
            cell.border = self.styles['border']
            if weekday_rate > 1.2:
                cell.font = Font(color='CC0000')
            elif weekday_rate < 0.5:
                cell.font = Font(color='0000CC')
            
            ws.cell(row=row, column=6, value=stats['假日值班']).border = self.styles['border']
            ws.cell(row=row, column=7, value=stats['假日配額']).border = self.styles['border']
            
            # 假日使用率
            holiday_rate = stats['假日值班'] / max(stats['假日配額'], 1)
            cell = ws.cell(row=row, column=8, value=f"{holiday_rate:.1%}")
            cell.border = self.styles['border']
            if holiday_rate > 1.2:
                cell.font = Font(color='CC0000')
            elif holiday_rate < 0.5:
                cell.font = Font(color='0000CC')
            
            ws.cell(row=row, column=9, value=stats['總值班數']).border = self.styles['border']
            
            row += 1
        
        # 調整欄寬
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_doctors_sheet(self, wb):
        """創建醫師清單工作表"""
        ws = wb.create_sheet(title="醫師清單")
        
        # 表頭
        headers = ['姓名', '角色', '平日配額', '假日配額', '不可值班日', '偏好值班日']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.styles['border']
        
        # 醫師資料
        row = 2
        for doctor in self.doctors:
            ws.cell(row=row, column=1, value=doctor.name).border = self.styles['border']
            ws.cell(row=row, column=2, value=doctor.role).border = self.styles['border']
            ws.cell(row=row, column=3, value=doctor.weekday_quota).border = self.styles['border']
            ws.cell(row=row, column=4, value=doctor.holiday_quota).border = self.styles['border']
            
            # 不可值班日（只顯示日期）
            unavailable_days = []
            for date_str in doctor.unavailable_dates:
                try:
                    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                    if date_obj.year == self.year and date_obj.month == self.month:
                        unavailable_days.append(str(date_obj.day))
                except:
                    pass
            ws.cell(row=row, column=5, value=', '.join(unavailable_days) if unavailable_days else '無').border = self.styles['border']
            
            # 偏好值班日（只顯示日期）
            preferred_days = []
            for date_str in doctor.preferred_dates:
                try:
                    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                    if date_obj.year == self.year and date_obj.month == self.month:
                        preferred_days.append(str(date_obj.day))
                except:
                    pass
            ws.cell(row=row, column=6, value=', '.join(preferred_days) if preferred_days else '無').border = self.styles['border']
            
            row += 1
        
        # 調整欄寬
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20